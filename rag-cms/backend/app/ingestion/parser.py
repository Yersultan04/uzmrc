from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import fitz  # PyMuPDF


@dataclass
class ParsedTable:
    """A table extracted from a page, already rendered as a markdown table."""

    page_number: int
    markdown: str
    rows: int
    cols: int
    # "vertical" = headers in the first row (classic); "horizontal" = headers in the
    # first column (transposed key/value style — common in product spec sheets).
    orientation: str = "vertical"
    # Labels extracted regardless of orientation — these feed the embedding so a
    # query like "первоначальный взнос" finds the table even when the description
    # doesn't quote that exact word.
    labels: list[str] = field(default_factory=list)


@dataclass
class ParsedPage:
    page_number: int  # 1-based
    text: str         # non-table text (table regions are subtracted, when detected)
    tables: list[ParsedTable] = field(default_factory=list)


# ---------------- PDF ----------------

# Tables smaller than this are usually noise (false positives from find_tables).
_PDF_TABLE_MIN_CELLS = 6


def _overlaps(a, b, threshold: float = 0.5) -> bool:
    """Returns True if a (bbox) is mostly inside b (bbox). Bboxes are (x0,y0,x1,y1)."""
    ix0 = max(a[0], b[0]); iy0 = max(a[1], b[1])
    ix1 = min(a[2], b[2]); iy1 = min(a[3], b[3])
    if ix0 >= ix1 or iy0 >= iy1:
        return False
    inter = (ix1 - ix0) * (iy1 - iy0)
    a_area = max(1.0, (a[2] - a[0]) * (a[3] - a[1]))
    return inter / a_area >= threshold


def _text_outside_bboxes(page, table_bboxes: list[tuple[float, float, float, float]]) -> str:
    """Reconstruct page text excluding text blocks that overlap any table bbox.
    Falls back to full page text if extraction fails."""
    try:
        d = page.get_text("dict")
    except Exception:
        return page.get_text("text") or ""
    parts: list[str] = []
    for block in d.get("blocks", []):
        if "lines" not in block:
            continue
        bbox = tuple(block.get("bbox", (0, 0, 0, 0)))
        if any(_overlaps(bbox, tb) for tb in table_bboxes):
            continue
        for line in block.get("lines", []):
            line_parts = []
            for span in line.get("spans", []):
                line_parts.append(span.get("text", ""))
            if line_parts:
                parts.append("".join(line_parts))
        parts.append("")
    return "\n".join(parts).strip()


_NUMERIC_RE = __import__("re").compile(r"^[\-\+]?[\d\s.,/%–\-]+$")


def _is_numeric_cell(s: str) -> bool:
    s = s.strip()
    if not s:
        return False
    return bool(_NUMERIC_RE.match(s))


def _text_ratio(cells: list[str]) -> float:
    """Fraction of non-empty cells that are textual (not numeric)."""
    non_empty = [c for c in cells if c.strip()]
    if not non_empty:
        return 0.0
    textual = sum(1 for c in non_empty if not _is_numeric_cell(c))
    return textual / len(non_empty)


def _detect_orientation(rows: list[list[str]]) -> str:
    """Return 'horizontal' if headers seem to be in the first COLUMN (transposed
    key/value style), else 'vertical' (headers in the first row)."""
    if len(rows) < 2 or len(rows[0]) < 2:
        return "vertical"
    first_row = rows[0]
    first_col = [r[0] for r in rows if r]
    row_text = _text_ratio(first_row)
    col_text = _text_ratio(first_col)
    # Strong signal: the first column is mostly text labels AND the first row has
    # significantly less text (i.e. data values bleed in).
    if col_text >= 0.8 and col_text - row_text >= 0.25:
        return "horizontal"
    return "vertical"


def _table_rows_to_markdown(rows: list[list[str]]) -> tuple[str, int, int, str, list[str]]:
    """Render a 2D list of cells to a markdown table.
    Returns (markdown, n_rows, n_cols, orientation, labels)."""
    cleaned: list[list[str]] = []
    for row in rows:
        cleaned.append([_cell_to_str(c) for c in row])
    if not cleaned:
        return "", 0, 0, "vertical", []
    width = max((len(r) for r in cleaned), default=0)
    if width == 0:
        return "", 0, 0, "vertical", []
    cleaned = [(r + [""] * width)[:width] for r in cleaned]

    orientation = _detect_orientation(cleaned)
    if orientation == "horizontal":
        # Labels are in the first column; column headers are missing/numeric.
        labels = [r[0] for r in cleaned if r and r[0].strip()]
        header = [f"col{i + 1}" if i > 0 else "Поле" for i in range(width)]
    else:
        header = cleaned[0]
        if not any(c.strip() for c in header):
            header = [f"c{i + 1}" for i in range(width)]
        labels = [c for c in header if c.strip()]
        cleaned = cleaned[1:]  # rest = body

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    for r in cleaned:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines), (len(cleaned) + 1), width, orientation, labels


def parse_pdf(path: Path) -> list[ParsedPage]:
    pages: list[ParsedPage] = []
    with fitz.open(path) as doc:
        for i, page in enumerate(doc, start=1):
            tables: list[ParsedTable] = []
            table_bboxes: list[tuple[float, float, float, float]] = []
            try:
                finder = page.find_tables()
                table_objs = getattr(finder, "tables", []) or []
            except Exception:
                table_objs = []
            for t in table_objs:
                try:
                    rows = t.extract()
                except Exception:
                    continue
                if not rows or sum(len(r) for r in rows) < _PDF_TABLE_MIN_CELLS:
                    continue
                md, n_rows, n_cols, orient, labels = _table_rows_to_markdown(rows)
                if not md:
                    continue
                tables.append(ParsedTable(
                    page_number=i, markdown=md, rows=n_rows, cols=n_cols,
                    orientation=orient, labels=labels,
                ))
                bbox = getattr(t, "bbox", None)
                if bbox is not None:
                    table_bboxes.append(tuple(bbox))

            if table_bboxes:
                text = _text_outside_bboxes(page, table_bboxes)
            else:
                text = (page.get_text("text") or "").strip()

            pages.append(ParsedPage(page_number=i, text=text, tables=tables))
    return pages


# ---------------- Plain text ----------------

def parse_text_file(path: Path) -> list[ParsedPage]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    return [ParsedPage(page_number=1, text=raw)]


# ---------------- XLSX ----------------

_XLSX_MAX_ROWS_PER_SHEET = 200      # raw rows shown per sheet before truncation
_XLSX_MAX_COLS_PER_SHEET = 40       # cap width to keep markdown manageable
_XLSX_MAX_CELL_LEN = 200            # truncate individual cells


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", " ").strip()
    if len(s) > _XLSX_MAX_CELL_LEN:
        s = s[: _XLSX_MAX_CELL_LEN - 1] + "…"
    # markdown table cell can't contain raw pipes
    return s.replace("|", "\\|")


def _rows_to_markdown(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    width = min(width, _XLSX_MAX_COLS_PER_SHEET)
    header = rows[0][:width]
    if not any(c.strip() for c in header):
        header = [f"c{i + 1}" for i in range(width)]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    for r in rows[1:]:
        padded = (r + [""] * width)[:width]
        lines.append("| " + " | ".join(padded) + " |")
    return "\n".join(lines)


def parse_xlsx(path: Path) -> list[ParsedPage]:
    """Each worksheet becomes its own ParsedPage carrying the sheet as a markdown TABLE
    (in `.tables`), with an empty `.text` body — the agent will see the table as a chunk
    via the table pipeline (LLM description embedding)."""
    from openpyxl import load_workbook

    pages: list[ParsedPage] = []
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        for idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            collected: list[list[str]] = []
            total_rows = 0
            for row in ws.iter_rows(values_only=True):
                total_rows += 1
                if len(collected) < _XLSX_MAX_ROWS_PER_SHEET:
                    collected.append([_cell_to_str(c) for c in row])
            head = f"[Sheet: {sheet_name}]"
            if total_rows > _XLSX_MAX_ROWS_PER_SHEET:
                head += (
                    f" (truncated: rows 1..{_XLSX_MAX_ROWS_PER_SHEET}"
                    f" of {total_rows} shown)"
                )
            if collected:
                md, n_rows, n_cols, orient, labels = _table_rows_to_markdown(collected)
                full_md = f"{head}\n\n{md}"
                tables = [ParsedTable(
                    page_number=idx, markdown=full_md,
                    rows=n_rows, cols=n_cols,
                    orientation=orient, labels=labels,
                )]
            else:
                tables = []
            pages.append(ParsedPage(page_number=idx, text="", tables=tables))
    finally:
        wb.close()
    return pages


# ---------------- Dispatch ----------------

def parse_file(path: Path, mime_type: str | None) -> list[ParsedPage]:
    suffix = path.suffix.lower()
    if suffix == ".pdf" or (mime_type and "pdf" in mime_type):
        return parse_pdf(path)
    if suffix == ".xlsx" or (
        mime_type and "spreadsheetml" in mime_type
    ):
        return parse_xlsx(path)
    if suffix in {".txt", ".md"} or (mime_type and mime_type.startswith("text/")):
        return parse_text_file(path)
    raise ValueError(f"Unsupported file type: suffix={suffix}, mime={mime_type}")
